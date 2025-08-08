#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ExcelファイルからLPラフ案を直接生成するスクリプト
"""

import os
import sys
import json
import openpyxl
from datetime import datetime
from typing import Dict, Any, List
from lp_rough_generator import LPRoughGenerator
from docbase_lp_uploader import DocbaseLPUploader

class ExcelToLPGenerator:
    def __init__(self):
        """初期化"""
        self.lp_generator = LPRoughGenerator()
        
    def parse_excel_to_json(self, excel_path: str) -> Dict[str, Any]:
        """ExcelファイルをJSON形式に変換"""
        
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        product_data = {}
        sku_list = []
        lp_structure = []
        page_details = []
        
        # セクションごとに処理
        current_section = "basic"
        page_data = {}
        
        # 全行を読み取り
        for row in sheet.iter_rows(values_only=True):
            if not row or not row[0]:  # 空行
                continue
                
            # セクション判定
            if str(row[0]).find("====") != -1:
                if "SKU" in str(row[0]):
                    current_section = "sku"
                    continue
                elif "LP構成" in str(row[0]):
                    current_section = "structure"
                    continue
                elif "各ページ詳細" in str(row[0]):
                    current_section = "page_details"
                    continue
                elif "その他" in str(row[0]):
                    current_section = "other"
                    continue
            
            # 基本情報
            if current_section == "basic":
                if row[0] and str(row[0]) != "項目名":
                    product_data[str(row[0])] = str(row[1]) if len(row) > 1 and row[1] else ""
            
            # SKU情報
            elif current_section == "sku":
                if row[0] and str(row[0]) != "sku_type":
                    if len(row) >= 3 and row[0] and row[1] and row[2]:
                        sku_list.append({
                            'type': str(row[0]),
                            'sku': str(row[1]),
                            'jan': str(row[2])
                        })
            
            # LP構成
            elif current_section == "structure":
                if row[0] and str(row[0]).startswith("page_"):
                    lp_structure.append(str(row[1]) if len(row) > 1 and row[1] else "")
            
            # ページ詳細
            elif current_section == "page_details":
                if row[0] and str(row[0]).startswith("page_"):
                    key = str(row[0])
                    value = str(row[1]) if len(row) > 1 and row[1] else ""
                    
                    # ページ番号を取得
                    parts = key.split("_")
                    if len(parts) >= 2:
                        try:
                            page_num = int(parts[1])
                            
                            # ページデータを初期化
                            if page_num not in page_data:
                                page_data[page_num] = {}
                            
                            # データタイプを判定
                            if "_text" in key:
                                page_data[page_num]['text'] = value.replace('\\n', '\n')
                            elif "_layout_note" in key:
                                page_data[page_num]['layout_note'] = value
                            elif "_image_" in key:
                                if 'images' not in page_data[page_num]:
                                    page_data[page_num]['images'] = []
                                if value:  # 画像URLがある場合のみ追加
                                    page_data[page_num]['images'].append(value)
                        except ValueError:
                            pass
            
            # その他の情報
            elif current_section == "other":
                if row[0] and len(row) > 1 and row[1]:
                    product_data[str(row[0])] = str(row[1]).replace('\\n', '\n')
        
        # ページ詳細を配列に変換
        for page_num in sorted(page_data.keys()):
            page_info = page_data[page_num]
            if not page_info.get('images'):
                page_info['has_images'] = True  # 画像準備中フラグ
            page_details.append(page_info)
        
        # 最終的なデータ構造を作成
        if sku_list:
            product_data['sku_list'] = sku_list
        if lp_structure:
            product_data['lp_structure'] = lp_structure
        if page_details:
            product_data['page_details'] = page_details
        
        return product_data
    
    def create_powerarq_excel_template(self, output_path: str = "templates/powerarq_blanket_lite.xlsx") -> str:
        """PowerArQ Electric Blanket LiteのExcelテンプレートを作成"""
        
        # ディレクトリ作成
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # 新しいワークブック作成
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "PowerArQ Electric Blanket Lite"
        
        # データ入力
        data = [
            # 基本情報セクション
            ["項目名", "内容"],
            ["product_name", "PowerArQ Electric Blanket Lite"],
            ["purpose", "PowerArQ Electric Blanket Liteの販売促進とブランド認知向上のため"],
            ["price", "9,000円（税込）"],
            ["release_date", "2025年10月上旬"],
            ["brand_value", "PowerARQブランドの信頼性と高品質"],
            ["target_users", "寒がりの方、電気代を節約したい方、テレワーカー、高齢者"],
            ["main_appeal", "10段階温度調節×タイマー機能で快適な温もりを"],
            [""],  # 空行
            
            # SKU情報セクション  
            ["==== SKU・JAN情報 ===="],
            ["sku_type", "SKU", "JAN"],
            ["グレー", "PAQ-BLANKET-LITE-GR", "4573211999999"],
            ["ベージュ", "PAQ-BLANKET-LITE-BG", "4573211999998"],
            [""],  # 空行
            
            # LP構成セクション
            ["==== LP構成 ===="],
            ["page_1", "TOPキャッチ・商品紹介"],
            ["page_2", "PowerARQブランドの信頼性"],
            ["page_3", "メイン機能・10段階温度調節"],
            ["page_4", "タイマー機能・省エネ性"],
            ["page_5", "使用シーン・ライフスタイル提案"],
            ["page_6", "サイズ・スペック詳細"],
            ["page_7", "カラーバリエーション"],
            ["page_8", "安全機能・品質保証"],
            ["page_9", "よくある質問"],
            ["page_10", "購入特典・キャンペーン情報"],
            [""],  # 空行
            
            # 各ページ詳細セクション
            ["==== 各ページ詳細 ===="],
            ["page_1_text", "PowerArQ Electric Blanket Lite\\n\\n心地よい温もりを、もっと身近に\\n\\n• 10段階の細かい温度調節\\n• 軽量2.0kgで持ち運び便利\\n• 1〜8時間タイマー機能"],
            ["page_1_layout_note", "商品の温かみを感じるメインビジュアル"],
            
            ["page_2_text", "PowerARQブランド\\n\\n信頼と品質の証\\n\\n累計販売台数○万台突破\\n※2025年○月時点"],
            ["page_2_layout_note", "ブランドロゴと実績を前面に"],
            
            ["page_3_text", "10段階の温度調節\\n\\nあなた好みの温かさを\\n\\n細かな調節で快適温度をキープ"],
            ["page_3_layout_note", "温度調節の操作イメージ"],
            
            ["page_4_text", "タイマー機能で省エネ\\n\\n1〜8時間の設定で\\n電気代を抑えながら快適に"],
            ["page_4_layout_note", "タイマー設定画面とコスト比較"],
            
            ["page_5_text", "いつでも、どこでも温かく\\n\\nリビング・寝室・書斎\\nあらゆるシーンで活躍"],
            ["page_5_layout_note", "様々な使用シーンの写真"],
            
            ["page_6_text", "仕様・スペック\\n\\nサイズ：188cm × 130cm\\n重量：2.0kg\\n消費電力：100V 115W"],
            ["page_6_layout_note", "スペック表とサイズ感の比較"],
            
            ["page_7_text", "カラーバリエーション\\n\\nグレー・ベージュの2色展開\\nお部屋に合わせてお選びください"],
            ["page_7_layout_note", "2色並べたカラー比較"],
            
            ["page_8_text", "安全機能搭載\\n\\n過熱防止機能で安心\\nPSEマーク取得済み"],
            ["page_8_layout_note", "安全認証マークと機能説明"],
            
            ["page_9_text", "よくある質問\\n\\nQ: 電気代はどのくらい？\\nA: 1時間あたり約○円（中温時）"],
            ["page_9_layout_note", "FAQ形式で見やすく"],
            
            ["page_10_text", "今なら特典付き\\n\\n送料無料\\n1年間の品質保証"],
            ["page_10_layout_note", "特典内容を目立たせて"],
            [""],  # 空行
            
            # その他情報セクション
            ["==== その他 ===="],
            ["specifications", "サイズ：188cm × 130cm\\n重量：2.0kg\\n消費電力：100V 115W\\n温度調節：10段階\\nタイマー：1〜8時間\\n安全機能：過熱防止機能"],
            ["features", "10段階の温度調節機能\\n1〜8時間のタイマー機能\\n過熱防止機能による安全性\\n軽量設計（2.0kg）\\n大判サイズ（188cm×130cm）"],
            ["usage_scenes", "リビング、寝室、書斎、オフィス"],
            ["design_variants", "グレー系、ベージュ系"]
        ]
        
        # データをシートに書き込み
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # 列幅調整
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 20
        
        # 保存
        workbook.save(output_path)
        return output_path
    
    def generate_lp_from_excel(self, excel_path: str, upload_to_docbase: bool = False) -> Dict[str, Any]:
        """ExcelファイルからLPラフ案を生成"""
        print(f"\n📊 Excel解析開始: {excel_path}")
        
        # Excelデータ解析
        try:
            product_data = self.parse_excel_to_json(excel_path)
            print(f"✅ Excel解析完了: {product_data.get('product_name', '商品名不明')}")
        except Exception as e:
            print(f"❌ Excel読み込みエラー: {e}")
            return None
        
        # LPラフ案生成
        lp_content = self.lp_generator.generate_lp_rough(product_data)
        
        # ファイル保存
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        product_name = product_data.get('product_name', '商品名').replace(' ', '_')
        output_filename = f"lp_rough_{product_name}_{timestamp}.md"
        output_path = os.path.join('output', output_filename)
        
        # outputディレクトリが存在しない場合は作成
        os.makedirs('output', exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(lp_content)
        
        print(f"✅ LPラフ案生成完了: {output_path}")
        
        result = {
            'lp_content': lp_content,
            'output_path': output_path,
            'product_data': product_data
        }
        
        # Docbaseアップロード
        if upload_to_docbase:
            try:
                uploader = DocbaseLPUploader()
                title = f"【LPラフ案】{product_data.get('product_name', '商品名')}"
                tags = ['LPラフ案', 'Excel生成', '自動生成']
                
                # 商品名でタグを追加
                if 'PowerArQ' in product_data.get('product_name', ''):
                    tags.extend(['PowerArQ', '電気毛布'])
                
                print(f"\n📤 Docbaseにアップロード中...")
                docbase_result = uploader.create_lp_post(title, lp_content, tags)
                
                result['docbase_url'] = docbase_result['url']
                result['docbase_id'] = docbase_result['id']
                
                print(f"✅ Docbaseアップロード完了")
                print(f"📎 URL: {docbase_result['url']}")
                
            except Exception as e:
                print(f"❌ Docbaseアップロードエラー: {e}")
        
        return result

def main():
    """メイン処理"""
    
    if len(sys.argv) < 2:
        print("使用方法:")
        print("  python excel_to_lp_generator.py <Excelファイル> [--upload]")
        print("  python excel_to_lp_generator.py --create-template")
        print("\n例:")
        print("  python excel_to_lp_generator.py templates/powerarq_blanket_lite.xlsx")
        print("  python excel_to_lp_generator.py templates/powerarq_blanket_lite.xlsx --upload")
        print("  python excel_to_lp_generator.py --create-template")
        sys.exit(1)
    
    # テンプレート作成モード
    if sys.argv[1] == '--create-template':
        try:
            generator = ExcelToLPGenerator()
            template_path = generator.create_powerarq_excel_template()
            print(f"✅ テンプレート作成完了: {template_path}")
            print("このファイルを編集して、python excel_to_lp_generator.py で使用してください")
        except Exception as e:
            print(f"❌ テンプレート作成エラー: {e}")
        sys.exit(0)
    
    excel_path = sys.argv[1]
    upload_flag = '--upload' in sys.argv
    
    if not os.path.exists(excel_path):
        print(f"❌ Excelファイルが見つかりません: {excel_path}")
        sys.exit(1)
    
    try:
        generator = ExcelToLPGenerator()
        result = generator.generate_lp_from_excel(excel_path, upload_to_docbase=upload_flag)
        
        if result:
            print(f"\n🎉 処理完了！")
            print(f"📁 出力ファイル: {result['output_path']}")
            if 'docbase_url' in result:
                print(f"🌐 Docbase URL: {result['docbase_url']}")
        else:
            print("❌ 処理に失敗しました")
            sys.exit(1)
            
    except Exception as e:
        print(f"❌ エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()